import openpyxl
from pathlib import Path
from schema import LegalDocumentExtraction

class ExcelExporter:
    """Exports structured legal data to the predefined Excel template."""
    
    def __init__(self, template_path: str = "output/Import Template.xlsx"):
        self.template_path = Path(template_path)
    
    def export(self, data: LegalDocumentExtraction) -> str:
        """Appends the data as a new row to the Excel file and saves it."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found at {self.template_path}")
            
        wb = openpyxl.load_workbook(self.template_path)
        sheet = wb.active
        
        # Read headers from first row
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}
        
        # Safely access properties with getattr to avoid any undefined None errors
        cp = data.case_instructing_party or type('dummy', (), {'agency_ref': None, 'agency_name': None, 'instructor_reference': None, 'instructing_party': None})()
        cd = data.claimant_details or type('dummy', (), {'title': '', 'forenames': '', 'surname': '', 'address': '', 'postcode': '', 'date_of_birth': '', 'telephone_home': '', 'telephone_work': '', 'mobile': '', 'email': '', 'medco_reference_no': '', 'occupation': ''})()
        ta = data.tracking_and_administration or type('dummy', (), {'accident_date': '', 'expert_name': '', 'source': ''})()
        gen = data.general or type('dummy', (), {'appointment_type': '', 'medical_date': '', 'appointment_time': '', 'consulting_venue_location': '', 'interpreter_required': '', 'interpreter_name': '', 'interpreter_company': '', 'interpreter_telephone': '', 'claims_handler_name': '', 'claims_handler_telephone': '', 'case_handler_email': ''})()
        acc = data.accident or type('dummy', (), {'accident_time': '', 'type_of_accident': '', 'accident_description': '', 'involvement': '', 'wearing_seatbelt': '', 'claimant_vehicle_reg': ''})()
        rp = data.responsible_parties or type('dummy', (), {'responsible_party_vehicle_reg': '', 'driver_first_name': '', 'driver_last_name': '', 'claimant_statement_of_responsibility': ''})()
        sp = data.special_instructions or type('dummy', (), {'injuries_symptoms': '', 'other_special_instructions': ''})()
        
        # Build client name
        client_name_parts = [cd.title, cd.forenames, cd.surname]
        client_name = " ".join([str(p) for p in client_name_parts if p]).strip()
        
        # Split address loosely
        address = str(cd.address or "")
        address_parts = [p.strip() for p in address.split(',') if p.strip()]
        c_address_1 = address_parts[0] if len(address_parts) > 0 else ""
        c_address_2 = address_parts[1] if len(address_parts) > 1 else ""
        c_address_3 = address_parts[2] if len(address_parts) > 2 else ""
        city = address_parts[-1] if len(address_parts) > 3 else ""

        # Construct the row map
        row_data = {
            "Agency Reference": cp.agency_ref,
            "Agency Name": cp.agency_name,
            "Instructor reference": cp.instructor_reference,
            "Instructor name": cp.instructing_party,
            "Client Name": client_name,
            "CAddress1": c_address_1,
            "CAddress2": c_address_2,
            "CAddress3": c_address_3,
            "City": city,
            "Postcode": cd.postcode,
            "Date Of Birth": cd.date_of_birth,
            "Date Of Accident": ta.accident_date,
            "Tel Home": cd.telephone_home,
            "Tel Work": cd.telephone_work,
            "Mobile": cd.mobile,
            "Client email": cd.email,
            "Appointment type": gen.appointment_type,
            "Dr Used": ta.expert_name,
            "Appointment Date": gen.medical_date,
            "Time": gen.appointment_time,
            "Venue1": ta.source or gen.consulting_venue_location,
            "MedcoRef": cd.medco_reference_no,
            "Interpreter required": str(gen.interpreter_required) if gen.interpreter_required is not None else "",
            "Interpreter Name": gen.interpreter_name,
            "Interpreter Company ": gen.interpreter_company, # Based on header exact match
            "Interpreter Telephone ": gen.interpreter_telephone,
            "Occupation": cd.occupation,
            "Claim Hander Name": gen.claims_handler_name,
            "Claim Hander Telephone": gen.claims_handler_telephone,
            "Claim Hander Email": gen.case_handler_email,
            "Accident Time": acc.accident_time,
            "Type of accident": acc.type_of_accident,
            "Accident Description": acc.accident_description,
            "Involvement": acc.involvement,
            "Seatbelt wearing": str(acc.wearing_seatbelt) if acc.wearing_seatbelt is not None else "",
            "Client vehicle reg number": acc.claimant_vehicle_reg,
            "Defendant party vehicle reg number": rp.responsible_party_vehicle_reg,
            "Defendant first name": rp.driver_first_name,
            "Defendant Last name": rp.driver_last_name,
            "Claimant version of accident": rp.claimant_statement_of_responsibility,
            "Injuries/Symptoms": sp.injuries_symptoms,
            "Any other special instructions": sp.other_special_instructions
        }
        
        # Append as new row
        new_row = []
        for header_idx in range(sheet.max_column):
            header_name = str(sheet.cell(row=1, column=header_idx+1).value)
            val = row_data.get(header_name, "")
            new_row.append(val)
            
        sheet.append(new_row)
        
        # Save modifications
        wb.save(self.template_path)
        return str(self.template_path)
